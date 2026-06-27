#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa questoes.xlsx para o banco SQLite do SimuladoPH.

Uso:
    python importar_questoes.py

Execute UMA VEZ antes de rodar a aplicação (ou quando atualizar a planilha).
O script limpa as tabelas de questões e reimporta tudo — não afeta usuários
nem histórico de simulados.
"""

import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Instale openpyxl: pip install openpyxl")

# Importa o app e os models para usar o contexto Flask
sys.path.insert(0, str(Path(__file__).parent))
from app import app, db
from models import Questao, Alternativa, SimuladoQuestao, Resposta

XLSX = Path(__file__).parent / "questoes.xlsx"
LETRAS_VALIDAS = {"A", "B", "C", "D"}


def normalizar(txt):
    if not txt:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(txt))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).strip()


def importar():
    if not XLSX.exists():
        sys.exit(f"Arquivo não encontrado: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Questões"] if "Questões" in wb.sheetnames else wb.active

    with app.app_context():
        db.create_all()

        # Limpa na ordem correta para respeitar FKs do PostgreSQL
        print("Limpando tabelas de questões...")
        Resposta.query.delete()
        SimuladoQuestao.query.delete()
        Alternativa.query.delete()
        Questao.query.delete()
        db.session.commit()

        inseridas = 0
        ignoradas = 0

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            enunciado = str(row[0]).strip() if row[0] else ""
            alt_a     = str(row[1]).strip() if row[1] else ""
            alt_b     = str(row[2]).strip() if row[2] else ""
            alt_c     = str(row[3]).strip() if row[3] else ""
            alt_d     = str(row[4]).strip() if row[4] else ""
            gabarito  = str(row[5]).strip().upper() if row[5] else ""
            explicacao = str(row[6]).strip() if row[6] else ""
            lista_ph  = str(row[7]).strip() if row[7] else ""
            disciplina = normalizar(row[8]) if row[8] else ""
            imagem    = str(row[9]).strip() if len(row) > 9 and row[9] else None
            turma     = str(row[10]).strip() if len(row) > 10 and row[10] else "7ano"
            if turma not in ("7ano", "8ano"):
                turma = "7ano"

            # Normaliza disciplina: garante acentuação correta em todas
            mapa_disciplinas = {
                "ciencias": "Ciências",
                "matematica": "Matemática",
                "historia": "História",
                "portugues": "Português",
                "geografia": "Geografia",
            }
            disciplina = mapa_disciplinas.get(disciplina.lower(), disciplina)

            if gabarito not in LETRAS_VALIDAS:
                print(f"  Linha {idx}: gabarito inválido ({row[5]!r}) — ignorada.")
                ignoradas += 1
                continue

            if imagem and imagem.upper() in ("X", "N/A", ""):
                imagem = None

            q = Questao(
                enunciado=enunciado,
                lista_ph=lista_ph,
                disciplina=disciplina,
                gabarito=gabarito,
                explicacao=explicacao,
                imagem=imagem,
                turma=turma,
            )
            db.session.add(q)
            db.session.flush()  # gera o q.id

            for letra, texto in [("A", alt_a), ("B", alt_b), ("C", alt_c), ("D", alt_d)]:
                db.session.add(Alternativa(questao_id=q.id, letra=letra, texto=texto))

            inseridas += 1

        db.session.commit()
        print(f"\nImportação concluída: {inseridas} questões inseridas, {ignoradas} ignoradas.")


if __name__ == "__main__":
    importar()
