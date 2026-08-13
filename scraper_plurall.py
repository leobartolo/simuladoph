#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scraper_plurall.py
Extrai questões das listas PH do Plurall e atualiza questoes.xlsx.
Imagens são salvas em static/imagens/ e referenciadas na planilha.

Dependências:
    pip install playwright openpyxl requests
    playwright install chromium

Uso:
    python scraper_plurall.py                  # scrape listas novas (não presentes no xlsx)
    python scraper_plurall.py --lista PH10     # força re-scrape de uma lista específica
    python scraper_plurall.py --inspecionar    # abre browser, pausa numa questão → dumpa HTML
"""

import argparse
import asyncio
import os
import re
import sys
import unicodedata

# Força UTF-8 no stdout para evitar UnicodeEncodeError no Windows (cp1252)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from pathlib import Path
from urllib.parse import urljoin, urlparse

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
except ImportError:
    sys.exit("Instale playwright:  pip install playwright && playwright install chromium")

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Instale openpyxl:  pip install openpyxl")

# ─────────────────────────────────────────────
# Configuração
# ─────────────────────────────────────────────
PLURALL_USUARIO = os.environ.get("PLURALL_USUARIO", "")
PLURALL_SENHA   = os.environ.get("PLURALL_SENHA",   "")
LOGIN_URL       = "https://atividades.plurall.net/"

BASE_DIR    = Path(__file__).parent
XLSX_PATH   = BASE_DIR / "questoes.xlsx"
IMG_DIR     = BASE_DIR / "static" / "imagens"

MAPA_DISC = {
    "ciencias":          "Ciências",
    "matematica":        "Matemática",
    "historia":          "História",
    "portugues":         "Português",
    "lingua portuguesa": "Português",
    "geografia":         "Geografia",
}


def norm_disc(s: str) -> str:
    s = unicodedata.normalize("NFKD", s.lower().strip())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return MAPA_DISC.get(s, s.title())



# ─────────────────────────────────────────────
# Login
# ─────────────────────────────────────────────
async def _dump_debug(page, motivo: str):
    try:
        await page.screenshot(path="debug_login_falha.png", full_page=True)
        Path("debug_login_falha.html").write_text(await page.content(), encoding="utf-8")
        print(f"  → {motivo}. URL atual: {page.url}")
        print("  → Dump salvo em debug_login_falha.png / debug_login_falha.html")
    except Exception as e_dump:
        print(f"  → Não foi possível salvar dump de diagnóstico: {e_dump}")


async def _clicar_botao_principal(page):
    """Clica no botão de ação do formulário (Continuar/Entrar). No Plurall
    nem sempre é um <button type="submit">, então cai para busca por texto."""
    try:
        await page.locator(
            'button[type="submit"]:visible, input[type="submit"]:visible'
        ).first.click(timeout=5_000)
        return
    except PWTimeout:
        pass

    await page.locator(
        'button:visible:has-text("Continuar"), button:visible:has-text("Entrar"), '
        'button:visible:has-text("Acessar"), button:visible:has-text("Login")'
    ).first.click(timeout=10_000)


async def fazer_login(page):
    if not PLURALL_USUARIO or not PLURALL_SENHA:
        raise RuntimeError(
            "Credenciais do Plurall não definidas. "
            "Use as variáveis de ambiente PLURALL_USUARIO e PLURALL_SENHA."
        )
    print("→ Acessando Plurall...")
    await page.goto(LOGIN_URL, wait_until="load")

    # Aguarda o campo de login aparecer (o Plurall pode redirecionar para SSO)
    try:
        await page.wait_for_selector("input", timeout=20_000)
    except PWTimeout:
        await _dump_debug(page, "Nenhum campo de input apareceu na página de login")
        raise RuntimeError("Página de login não carregou — verifique a URL e a conexão.")

    # Tenta preencher o formulário. O Plurall usa login/username/text
    campo_login = page.locator(
        'input[name="login"], input[name="username"], input[name="user"], '
        'input[type="text"]:visible, input[name="email"]:visible'
    ).first
    await campo_login.fill(PLURALL_USUARIO)

    # Fecha banner de cookies se aparecer, para não atrapalhar cliques
    try:
        botao_cookies = page.get_by_text("Aceitar e fechar", exact=False)
        await botao_cookies.wait_for(state="visible", timeout=3_000)
        await botao_cookies.click()
    except PWTimeout:
        pass

    # Login em duas etapas: usuário primeiro, "Continuar", só então aparece a senha
    campo_senha = page.locator('input[type="password"]:visible').first
    senha_ja_visivel = True
    try:
        await campo_senha.wait_for(state="visible", timeout=1_500)
    except PWTimeout:
        senha_ja_visivel = False

    if not senha_ja_visivel:
        try:
            await _clicar_botao_principal(page)
        except PWTimeout:
            await _dump_debug(page, "Não achei o botão 'Continuar' da 1ª etapa")
            raise

    try:
        await campo_senha.wait_for(state="visible", timeout=20_000)
        await campo_senha.fill(PLURALL_SENHA)
    except PWTimeout:
        await _dump_debug(page, "Falha ao achar campo de senha")
        raise

    try:
        await _clicar_botao_principal(page)
    except PWTimeout:
        await _dump_debug(page, "Não achei o botão de login final")
        raise

    # Aguarda chegar em atividades.plurall.net (OAuth pode ter vários redirects)
    try:
        await page.wait_for_url("*atividades.plurall.net*", timeout=25_000)
    except PWTimeout:
        pass
    await page.wait_for_load_state("load", timeout=20_000)
    await page.wait_for_timeout(2000)  # tempo para o SPA renderizar
    print(f"  Login OK — URL: {page.url}")


# ─────────────────────────────────────────────
# Navegação: listas PH na aba "Estudo Orientado"
# ─────────────────────────────────────────────
async def ir_para_estudo_orientado(page):
    """Navega para atividades.plurall.net e abre a aba Estudo Orientado completa."""
    # Navega para a raiz a menos que já esteja nela (evita colisão com redirects OAuth)
    raiz = "https://atividades.plurall.net"
    if page.url.rstrip("/") != raiz:
        await page.goto(raiz + "/", wait_until="load")
    await page.wait_for_timeout(3000)
    print(f"  URL pós-navegação: {page.url}")

    # Se caiu na página de login, a sessão não foi estabelecida
    if "login.plurall" in page.url or "login" in page.url.split("plurall.net")[-1]:
        raise RuntimeError(
            f"Sessão não estabelecida — redirecionado para login: {page.url}\n"
            "Verifique se as credenciais estão corretas e se a conta usa login direto (não SSO escolar)."
        )

    # Descobre quais abas estão disponíveis para diagnóstico
    abas_disponiveis = await page.evaluate("""() =>
        Array.from(document.querySelectorAll('[data-test-id^="tab-"]'))
             .map(e => e.getAttribute('data-test-id'))
    """)
    print(f"  Abas visíveis: {abas_disponiveis}")

    # Tenta encontrar a aba PH (aceita variações de nome)
    aba_seletor = None
    for candidato in abas_disponiveis:
        if "ph" in candidato.lower() or "estudo-orientado" in candidato.lower():
            aba_seletor = f'[data-test-id="{candidato}"]'
            print(f"  Usando aba: {candidato}")
            break

    if not aba_seletor:
        # Fallback: aguarda até 30s pelo seletor original
        aba_seletor = '[data-test-id="tab-Estudo-Orientado-(pHs)"]'
        try:
            await page.wait_for_function(
                f"() => !!document.querySelector('{aba_seletor}')",
                timeout=30_000
            )
            print("  Aba encontrada após espera")
        except PWTimeout:
            raise RuntimeError(
                f"Aba 'Estudo Orientado (pHs)' não encontrada. "
                f"Abas disponíveis: {abas_disponiveis}\n"
                "Use --inspecionar para ver o que aparece na página."
            )

    # Clica na aba
    await page.evaluate(f"""() => {{
        const tab = document.querySelector('{aba_seletor}');
        if (tab) tab.click();
    }}""")

    # Aguarda os cards aparecerem (não usa wait fixo)
    try:
        await page.wait_for_selector('[data-test-id*="inactive-material"]', timeout=20_000)
    except PWTimeout:
        await page.wait_for_timeout(3000)

    n = await page.locator('[data-test-id*="inactive-material"]').count()
    print(f"  Cards encerrados visíveis: {n}")


async def listar_listas_ph(page) -> list[str]:
    """Retorna lista de lista_ids encerrados (ex: ['PH1','PH2',...])."""
    # Encerrado = cover do card tem data-test-id com "inactive-material"
    listas = []
    covers = page.locator('[data-test-id*="inactive-material"]')
    count = await covers.count()

    for i in range(count):
        titulo = await covers.nth(i).evaluate("""el => {
            const holder = el.closest('[class*="holder"]') || el.parentElement;
            return holder?.querySelector('[class*="title"]')?.innerText?.trim() || '';
        }""")
        m = re.search(r"pH\s*0*(\d+)", titulo, re.I)
        if not m:
            continue
        lista_id = f"PH{int(m.group(1)):02d}"
        listas.append(lista_id)
        print(f"  Lista encerrada: {lista_id}  ({titulo})")

    return listas


async def ir_para_lista_ph(page, lista_id: str) -> bool:
    """Navega para a página de uma lista PH clicando em Acessar no card."""
    await ir_para_estudo_orientado(page)

    num = lista_id.replace("PH", "").lstrip("0") or "0"
    ph_re = re.compile(rf"pH\s*0*{num}\b", re.I)

    covers = page.locator('[data-test-id*="inactive-material"]')
    count = await covers.count()

    for i in range(count):
        titulo = await covers.nth(i).evaluate("""el => {
            const holder = el.closest('[class*="holder"]') || el.parentElement;
            return holder?.querySelector('[class*="title"]')?.innerText?.trim() || '';
        }""")
        if not ph_re.search(titulo):
            continue

        # Clica em Acessar via JavaScript (o <a> não tem href)
        clicked = await covers.nth(i).evaluate("""el => {
            const holder = el.closest('[class*="holder"]') || el.parentElement;
            const btn = holder?.querySelector('[data-test-id="material-card-button-wrapper"] a');
            if (btn) { btn.click(); return true; }
            return false;
        }""")
        if clicked:
            await page.wait_for_load_state("load")
            await page.wait_for_timeout(2500)
            return True

    print(f"  ⚠ Card {lista_id} não encontrado ou não encerrado")
    return False


# ─────────────────────────────────────────────
# Navegação: disciplinas dentro de uma lista
# ─────────────────────────────────────────────
async def listar_disciplinas(page) -> list[dict]:
    """Na página de uma lista PH, retorna [{disciplina, href}].

    Estrutura confirmada pelo HTML:
      <a href="/material/.../tarefa/XXXXX">
        <div data-test-id="progress-group-task-title-N">
          <p>Ciências</p>
        </div>
        ...barra de progresso...
      </a>
    """
    await page.wait_for_timeout(2000)

    disciplinas = []
    # Links que contêm um progress-group-task-title (= cada disciplina)
    links = page.locator('a[href]:has([data-test-id*="progress-group-task-title"])')
    count = await links.count()

    for i in range(count):
        link = links.nth(i)
        href = await link.get_attribute("href")
        if not href:
            continue

        # Título está no <p> dentro do task-title
        titulo_el = link.locator('[data-test-id*="progress-group-task-title"] p').first
        if await titulo_el.count():
            titulo = (await titulo_el.inner_text()).strip()
        else:
            titulo = (await link.inner_text()).strip().split("\n")[0].strip()

        disciplinas.append({"disciplina": norm_disc(titulo), "href": href})

    return disciplinas


# ─────────────────────────────────────────────
# Extração: questão individual
# ─────────────────────────────────────────────
async def extrair_questao_atual(page, lista_id: str, disciplina: str, num: int, turma: str = "7ano") -> dict | None:
    """Extrai enunciado, alternativas, gabarito, explicação e imagem da questão visível."""
    await page.wait_for_load_state("load")
    await page.wait_for_timeout(1000)

    print(f"    [{num}] {page.url}")

    # ── Enunciado + imagens inline ────────────────────────────────────────────
    # Processa o innerHTML: cada <img> é baixada e substituída por [img:filename]
    # no texto. Isso preserva notações matemáticas inline (∠, △ etc.) e figuras.
    enunciado = ""
    imagem_filename = ""  # mantido por compatibilidade com questões antigas

    el = page.locator('[class*="question-text"]').first
    if await el.count():
        img_prefix = f"{turma[:2]}_" if turma != "7ano" else ""
        num_lista = lista_id.replace("PH", "").zfill(2)
        disc_raw = norm_disc(disciplina)[:3].lower()
        disc_abrev = "".join(
            c for c in unicodedata.normalize("NFKD", disc_raw)
            if not unicodedata.combining(c)
        )

        # Substitui cada <img> por um marcador de texto e coleta os srcs
        imgs_info = await page.evaluate("""() => {
            const el = document.querySelector('[class*="question-text"]');
            if (!el) return {text: '', imgs: []};
            const clone = el.cloneNode(true);
            const imgs = [];
            let idx = 0;
            clone.querySelectorAll('img').forEach(img => {
                const src = img.getAttribute('src') || '';
                imgs.push({idx, src});
                img.replaceWith(new Text('<<IMG' + idx + '>>'));
                idx++;
            });
            return {text: (clone.innerText || clone.textContent).trim(), imgs};
        }""")

        raw_text = imgs_info["text"]

        for img_info in imgs_info["imgs"]:
            src   = img_info["src"]
            idx   = img_info["idx"]
            marker = f"<<IMG{idx}>>"

            if not src or src.startswith("data:"):
                raw_text = raw_text.replace(marker, "")
                continue

            parsed_path = urlparse(src).path
            ext = Path(parsed_path).suffix.lower()
            if ext not in (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"):
                ext = ".png"

            suffix = f"_t{idx}" if idx > 0 else ""
            filename = f"{img_prefix}ph{num_lista}_{disc_abrev}_{num:02d}{suffix}{ext}"
            dest = IMG_DIR / filename
            IMG_DIR.mkdir(parents=True, exist_ok=True)

            try:
                response = await asyncio.wait_for(page.request.get(src), timeout=15)
                body = await response.body()
                if len(body) >= 100:
                    dest.write_bytes(body)
                    raw_text = raw_text.replace(marker, f"[img:{filename}]")
                    print(f"      → img {idx}: {filename} ({len(body)} bytes)")
                else:
                    raw_text = raw_text.replace(marker, "")
                    print(f"      ⚠ img {idx} vazia ({len(body)} bytes), ignorando")
            except Exception as e:
                raw_text = raw_text.replace(marker, "")
                print(f"      ⚠ erro img {idx}: {e}")

        enunciado = raw_text

    # ── Alternativas e Gabarito ───────────────────────────────────────────────
    # Estrutura confirmada pelo HTML:
    #   <li class="option">
    #     <span data-test-id="option" class="...Option-module_right__... ou ...wrong__...">
    #       <div class="...right-answer__... ou ...wrong-answer__...">
    #         <p>b</p>   ← letra em minúsculo
    #       </div>
    #     </span>
    #     <div class="...option-div-text__...">
    #       <span class="...option-text__..."><p>Texto da alternativa</p></span>
    #     </div>
    #   </li>
    alts: dict[str, str] = {}
    gabarito = ""

    for item in await page.locator("li.option").all():
        # Círculo colorido: right-answer (verde=correto) ou wrong-answer (vermelho=errado)
        circulo = item.locator('[class*="right-answer"], [class*="wrong-answer"]').first
        if not await circulo.count():
            continue

        # Letra dentro do círculo (vem em minúsculo: a, b, c, d)
        letra_el = circulo.locator("p").first
        if not await letra_el.count():
            continue
        letra = (await letra_el.inner_text()).strip().upper()
        if letra not in ("A", "B", "C", "D", "E"):
            continue

        # Texto da alternativa
        texto_el = item.locator('[class*="option-text"]').first
        texto = ""
        if await texto_el.count():
            texto = (await texto_el.inner_text()).strip()
            texto = re.sub(r"Alternativa selecionada\s*", "", texto, flags=re.I).strip()

        # Alternativa vazia → provavelmente é uma imagem (fórmula matemática etc.)
        if not texto:
            img_alt = item.locator('[class*="option-text"] img, [class*="option-div-text"] img').first
            if await img_alt.count():
                src = await img_alt.get_attribute("src") or ""
                if src and not src.startswith("data:"):
                    parsed_path = urlparse(src).path
                    ext = Path(parsed_path).suffix.lower()
                    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"):
                        ext = ".png"
                    img_prefix = f"{turma[:2]}_" if turma != "7ano" else ""
                    num_lista = lista_id.replace("PH", "").zfill(2)
                    disc_raw = norm_disc(disciplina)[:3].lower()
                    disc_abrev = "".join(
                        c for c in unicodedata.normalize("NFKD", disc_raw)
                        if not unicodedata.combining(c)
                    )
                    filename = f"{img_prefix}ph{num_lista}_{disc_abrev}_{num:02d}_{letra.lower()}{ext}"
                    dest = IMG_DIR / filename
                    IMG_DIR.mkdir(parents=True, exist_ok=True)
                    try:
                        response = await asyncio.wait_for(
                            page.request.get(src), timeout=15
                        )
                        body = await response.body()
                        if len(body) >= 100:
                            dest.write_bytes(body)
                            texto = f"[img:{filename}]"
                            print(f"      → alt {letra} como imagem: {filename}")
                    except Exception as e:
                        print(f"      ⚠ erro imagem alt {letra}: {e}")

        alts[letra] = texto

        # Gabarito: círculo com classe "right-answer" (confirmado no HTML)
        circulo_class = await circulo.get_attribute("class") or ""
        if "right-answer" in circulo_class:
            gabarito = letra

    if not alts:
        print(f"      ⚠ alternativas não encontradas — verifique os seletores")
    if not gabarito:
        print(f"      ⚠ gabarito não detectado (alts={list(alts)})")

    if not alts or not gabarito:
        return None

    # ── Explicação ───────────────────────────────────────────────────────────
    # Confirmado no HTML: class contém "feedback" dentro do answer-container
    explicacao = ""
    exp_el = page.locator('[class*="feedback"]').first
    if await exp_el.count():
        txt = (await exp_el.inner_text()).strip()
        explicacao = re.sub(r"^gabarito\s*", "", txt, flags=re.I).strip()

    return {
        "enunciado": enunciado,
        "A": alts.get("A", ""),
        "B": alts.get("B", ""),
        "C": alts.get("C", ""),
        "D": alts.get("D", ""),
        "E": alts.get("E", ""),
        "gabarito":   gabarito,
        "explicacao": explicacao,
        "lista":      lista_id,
        "disciplina": norm_disc(disciplina),
        "imagem":     imagem_filename,
        "turma":      turma,
    }


# ─────────────────────────────────────────────
# Scrape de uma disciplina completa
# ─────────────────────────────────────────────
DISCIPLINAS_IGNORAR = {"producao de texto e redacao", "producao de texto", "redacao"}


async def scrape_disciplina(page, lista_id: str, disciplina: str, tarefa_url: str | None, turma: str = "7ano") -> list[dict]:
    # Produção de Texto é dissertativa — sem gabarito A/B/C/D
    if norm_disc(disciplina).lower() in DISCIPLINAS_IGNORAR:
        return []

    url_base = "https://atividades.plurall.net"
    if tarefa_url:
        await page.goto(urljoin(url_base, tarefa_url), wait_until="load")
        await page.wait_for_timeout(3000)

    # Aguarda os tiles de exercício aparecerem; se falhar, recarrega e tenta de novo
    try:
        await page.wait_for_selector('a[href*="/exercicio/"]', timeout=25_000)
    except PWTimeout:
        print(f"  ⚠ tiles não apareceram — recarregando...")
        if tarefa_url:
            await page.goto(urljoin(url_base, tarefa_url), wait_until="load")
            await page.wait_for_timeout(5000)
        try:
            await page.wait_for_selector('a[href*="/exercicio/"]', timeout=25_000)
        except PWTimeout:
            print(f"  ⚠ tiles de exercício não apareceram para {disciplina}")
            return []

    try:
        primeiro = page.locator('a[href*="/exercicio/"]').first
        if await primeiro.count():
            await primeiro.click()
            await page.wait_for_load_state("load")
            await page.wait_for_timeout(2000)
    except PWTimeout:
        pass

    questoes = []
    num = 1

    while True:
        try:
            q = await extrair_questao_atual(page, lista_id, disciplina, num, turma)
            if q:
                questoes.append(q)
        except Exception as e:
            print(f"      ⚠ erro extraindo Q{num}: {e}")

        # Botão "próxima" confirmado pelo HTML: data-test-id="icon-Chevron-right"
        try:
            proxima = page.locator('[data-test-id="icon-Chevron-right"]').first
            if not await proxima.count():
                break

            url_antes = page.url
            await proxima.click()
            await page.wait_for_load_state("load")
            await page.wait_for_timeout(1500)

            if page.url == url_antes:
                break
        except Exception as e:
            print(f"      ⚠ erro navegando para Q{num+1}: {e}")
            break

        num += 1

    return questoes


# ─────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────
def _col_lista(ws) -> int:
    """Índice 0-based da coluna 'Lista' no Excel."""
    header = [ws.cell(1, c).value for c in range(1, 14)]
    for i, h in enumerate(header):
        if str(h or "").strip().upper() == "LISTA":
            return i
    tem_e = len(header) > 5 and str(header[5] or "").strip().upper() == "E"
    return 8 if tem_e else 7


def _col_turma(ws) -> int:
    """Índice 0-based da coluna 'Turma' no Excel."""
    header = [ws.cell(1, c).value for c in range(1, 14)]
    for i, h in enumerate(header):
        if str(h or "").strip().upper() == "TURMA":
            return i
    tem_e = len(header) > 5 and str(header[5] or "").strip().upper() == "E"
    return 11 if tem_e else 10


def listas_no_excel(turma: str | None = None) -> set[str]:
    """Retorna listas já presentes no Excel, opcionalmente filtradas por turma."""
    if not XLSX_PATH.exists():
        return set()
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Questões"] if "Questões" in wb.sheetnames else wb.active
    col_l = _col_lista(ws)
    col_t = _col_turma(ws)
    result = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= col_l or not row[col_l]:
            continue
        if turma and (len(row) <= col_t or str(row[col_t] or "").strip() != turma):
            continue
        result.add(str(row[col_l]).strip())
    return result


def remover_lista_excel(lista_id: str, turma: str | None = None):
    """Remove linhas de uma lista do Excel, respeitando a turma se informada."""
    if not XLSX_PATH.exists():
        return
    wb = load_workbook(XLSX_PATH)
    ws = wb["Questões"] if "Questões" in wb.sheetnames else wb.active
    col_l = _col_lista(ws)
    col_t = _col_turma(ws)
    linhas_remover = []
    for row in ws.iter_rows(min_row=2):
        val_lista = row[col_l].value if len(row) > col_l else None
        if not val_lista or str(val_lista).strip() != lista_id:
            continue
        if turma:
            val_turma = row[col_t].value if len(row) > col_t else None
            if str(val_turma or "").strip() != turma:
                continue
        linhas_remover.append(row[0].row)
    for row_num in reversed(linhas_remover):
        ws.delete_rows(row_num)
    wb.save(XLSX_PATH)
    turma_desc = f" ({turma})" if turma else ""
    print(f"  ✓ {len(linhas_remover)} linhas de {lista_id}{turma_desc} removidas do Excel")


def salvar_excel(novas: list[dict]):
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb["Questões"] if "Questões" in wb.sheetnames else wb.active
        # Migrar Excel antigo (sem coluna E) inserindo a coluna na posição 6
        header_row = [ws.cell(1, c).value for c in range(1, 8)]
        if len(header_row) >= 6 and header_row[5] != "E":
            ws.insert_cols(6)
            ws.cell(1, 6).value = "E"
            print("✓ Migração automática: coluna E adicionada ao Excel")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questões"
        ws.append(["Questao", "A", "B", "C", "D", "E", "Gabarito", "Explicacao", "Lista", "Disciplina", "Imagem", "Turma"])

    for q in novas:
        ws.append([
            q["enunciado"], q["A"], q["B"], q["C"], q["D"], q.get("E", ""),
            q["gabarito"], q["explicacao"], q["lista"], q["disciplina"],
            q["imagem"], q.get("turma", "7ano"),
        ])

    wb.save(XLSX_PATH)
    print(f"\n✓ {len(novas)} questões adicionadas → {XLSX_PATH}")


# ─────────────────────────────────────────────
# Modo Inspecionar
# ─────────────────────────────────────────────
async def modo_inspecionar(page):
    """
    Navega direto para uma questão com gabarito liberado e salva o HTML.
    URL hardcoded da questão 01 de Ciências do PH10 (visível nos prints).
    """
    # URL da questão 01 de Ciências do PH10 (dos prints enviados)
    URL_QUESTAO = (
        "https://atividades.plurall.net/material/13350536/aula/13350536"
        "/tarefa/12232920/exercicio/12233794/"
    )

    print("\n=== MODO INSPECIONAR ===")

    # 1) Salva HTML da página Estudo Orientado (cards PH)
    print("1) Salvando HTML da página principal (Estudo Orientado)...")
    await ir_para_estudo_orientado(page)
    html_main = await page.content()
    Path("debug_main.html").write_text(html_main, encoding="utf-8")
    print(f"   → debug_main.html ({len(html_main)} bytes)")

    # 2) Navega para pH10 e salva HTML da página de disciplinas
    print("2) Navegando para pH10 (clicando Acessar)...")
    ok = await ir_para_lista_ph(page, "PH10")
    if ok:
        html_ph = await page.content()
        Path("debug_ph10.html").write_text(html_ph, encoding="utf-8")
        print(f"   → debug_ph10.html ({len(html_ph)} bytes)  URL: {page.url}")
    else:
        print("   ⚠ não conseguiu navegar para pH10")

    # 3) Salva HTML de uma questão com gabarito
    print("3) Navegando para questão de exemplo...")
    await page.goto(URL_QUESTAO, wait_until="load")
    try:
        await page.wait_for_selector("li.option", timeout=15_000)
        await page.wait_for_timeout(2000)
    except PWTimeout:
        print("   ⚠ timeout aguardando conteúdo")
    html_q = await page.content()
    Path("debug_questao.html").write_text(html_q, encoding="utf-8")
    print(f"   → debug_questao.html ({len(html_q)} bytes)")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
async def main(lista_filtro: list[str] | None, inspecionar: bool, forcar: bool, turma: str = "7ano"):
    async with async_playwright() as pw:
        headless = not inspecionar  # headed só no modo --inspecionar (local)
        browser = await pw.chromium.launch(
            headless=headless,
            args=[
                "--disable-dev-shm-usage",   # usa /tmp ao invés de /dev/shm (padrão pequeno em containers)
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-gpu",
                "--disable-extensions",
                "--disable-background-networking",
                "--disable-default-apps",
                "--mute-audio",
                "--no-first-run",
            ],
        )
        context = await browser.new_context()
        page = await context.new_page()

        await fazer_login(page)

        if inspecionar:
            await modo_inspecionar(page)
            await browser.close()
            return

        await ir_para_estudo_orientado(page)
        listas_ph = await listar_listas_ph(page)

        if not listas_ph:
            print("Nenhuma lista PH encerrada encontrada. Verifique os seletores com --inspecionar.")
            await browser.close()
            return

        ja_no_excel = listas_no_excel(turma)
        print(f"\nListas já no Excel ({turma}): {ja_no_excel or '(nenhuma)'}")

        filtro_set = {l.upper() for l in lista_filtro} if lista_filtro else None

        # --forcar: remove do Excel para reextrair
        if forcar and filtro_set:
            for lista_up in filtro_set:
                if lista_up in ja_no_excel:
                    remover_lista_excel(lista_up, turma)
                    ja_no_excel.discard(lista_up)

        total_salvo = 0

        for lista_id in listas_ph:
            if filtro_set and lista_id.upper() not in filtro_set:
                continue
            if lista_id in ja_no_excel and not filtro_set:
                print(f"  {lista_id}: já no Excel, pulando.")
                continue

            print(f"\n── {lista_id} ──")
            try:
                ok = await ir_para_lista_ph(page, lista_id)
                if not ok:
                    continue

                disciplinas = await listar_disciplinas(page)
                if not disciplinas:
                    print(f"  ⚠ nenhuma disciplina encontrada para {lista_id}")
                    continue

                questoes_lista: list[dict] = []
                for disc in disciplinas:
                    print(f"  {disc['disciplina']} ({disc['href']})")
                    qs = await scrape_disciplina(page, lista_id, disc["disciplina"], disc["href"], turma)
                    print(f"  → {len(qs)} questões extraídas")
                    questoes_lista.extend(qs)

                if questoes_lista:
                    salvar_excel(questoes_lista)
                    total_salvo += len(questoes_lista)

            except Exception as e:
                print(f"  ⚠ erro na lista {lista_id}: {e}")
                print(f"  Continuando para próxima lista...")

        if total_salvo:
            print(f"\n✓ Total: {total_salvo} questões salvas no Excel")
            print("Próximo passo: python importar_questoes.py")
        else:
            print("\nNenhuma questão nova para importar.")

        await browser.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scraper das listas PH do Plurall")
    parser.add_argument("--lista", nargs="+", help="Filtra por uma ou mais listas (ex: PH10 ou PH1 PH2 PH3)")
    parser.add_argument("--forcar", action="store_true",
                        help="Remove a lista do Excel antes de reextrair (usar com --lista)")
    parser.add_argument("--inspecionar", action="store_true",
                        help="Abre browser pausado para inspecionar HTML")
    parser.add_argument("--turma", default=None, choices=["7ano", "8ano"],
                        help="Turma das questões a importar (7ano ou 8ano)")
    args = parser.parse_args()

    turma = args.turma or os.environ.get("SCRAPER_TURMA", "7ano")
    asyncio.run(main(args.lista or None, args.inspecionar, args.forcar, turma))
